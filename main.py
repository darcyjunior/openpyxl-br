# pip install openpyxl

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class Planilhas():

    # def __init__(self, caminho_arquivo: str):
    #     self.wb = load_workbook(caminho_arquivo)

    def abrir_arquivo_excel(self, caminho_arquivo: str):
        self.wb = load_workbook(caminho_arquivo)

    def fechar_arquivo_excel(self):
        self.wb.close()

    def salvar_arquivo_excel(self, nome_arquivo: str):
        self.wb.save(f'{nome_arquivo}.xlsx')

    def criar_planilhas(self, lista_nomes_planilhas: list):
        for planilha in lista_nomes_planilhas:
            self.wb.create_sheet(planilha)

    def apagar_planilha_por_nome(self, nome_planilha: str):
        self.wb.remove(self.wb[nome_planilha])

    def copiar_planilha(self, nome_planilha_para_copia: str, nome_nova_planilha=''):
        if nome_nova_planilha != '':
            planilha = self.wb[nome_planilha_para_copia]
            nova_planilha = self.wb.copy_worksheet(planilha)
            nova_planilha.title = nome_nova_planilha

    def buscar_planilha_por_indice(self, indice: int):
        try:
            print(self.wb.worksheets[indice])
            return self.wb.worksheets[indice]
        except IndexError:
            print(f'Planilha não encontrada para o índice: {indice}')

    def preenche_linhas_da_planilha_por_lista(self, nome_planilha: str, lista_dados: list):
        plan = self.wb[nome_planilha]
        for x, dados in enumerate(lista_dados):
            for y, item in enumerate(dados):
                plan.cell(row=x+1, column=y+1, value=item)

    def apaga_linhas_planilha(self, nome_planilha: str, linha_inicio: int, qtd_linhas: int):
        planilha = self.wb[nome_planilha]
        planilha.delete_rows(linha_inicio, qtd_linhas)

    def apaga_colunas_planilha(self, nome_planilha: str, col_inicio: int, qtd_cols: int):
        planilha = self.wb[nome_planilha]
        planilha.delete_cols(col_inicio, qtd_cols)

    def adiciona_linhas_por_dataframe(self, nome_planilha: str, df, index_df=False, header_df=False):
        planilha = self.wb[nome_planilha]
        # Faz o loop adicionando as linhas do dataframe
        for row in dataframe_to_rows(df, index=index_df, header=header_df):
            planilha.append(row)

    # Formatando as células

    def redimensiona_largura_colunas(self, nome_planilha: str, lista_colunas: list, largura: int):
        planilha = self.wb[nome_planilha]
        for coluna in lista_colunas:
            planilha.column_dimensions[coluna].width = largura

    def formata_moeda_colunas(self, nome_planilha: str, lista_colunas: list):
        planilha = self.wb[nome_planilha]
        for col in lista_colunas:
            for cell in planilha[col]:
                planilha[col + str(cell.row)
                         ].number_format = 'R$ #,##0.00;-R$ #,##0.00'

    def formata_data_colunas(self, nome_planilha: str, lista_colunas: list, formato_data='dd/mm/yyyy;@'):
        planilha = self.wb[nome_planilha]
        for col in lista_colunas:
            for cell in planilha[col]:
                planilha[col +
                         str(cell.row)].number_format = formato_data

    def quebra_texto_por_colunas(self, nome_planilha: str, lista_colunas: list):
        planilha = self.wb[nome_planilha]
        for col in lista_colunas:
            for cell in planilha[col]:
                conteudo = str(cell.value)
                cell.alignment = Alignment(
                    vertical='center', horizontal='center', wrap_text=True)
                # Remove espaço e quebras de linhas desnecessários
                conteudo.strip()


plan = Planilhas()
plan.abrir_arquivo_excel('nomedoarquivo.xlsx')

plan.formata_moeda_colunas('ATIVOS', ['I', 'R', 'S', 'T'])
plan.quebra_texto_por_colunas('ATIVOS', ['F', 'M'])
plan.redimensiona_largura_colunas('ATIVOS', ['A'], 100)
plan.salvar_arquivo_excel('nomedoarquivo2')
plan.fechar_arquivo_excel()
